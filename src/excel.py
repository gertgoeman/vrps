import openpyxl
import datetime

class RecordSet(object):
    def __init__(self, columns, entries):
        self.columns = columns
        self.entries = entries

def read_locations(path):
    wb = openpyxl.load_workbook(filename = path)
    ws = wb.active

    first_row = next(ws.rows)
    columns = [ cell.value or "" for cell in first_row ]

    entries = []

    for i, row in enumerate(ws.rows):
        if i == 0: continue     # the first row contains the headers
        entry = {}
        for index, cell in enumerate(row):
            column = ""
            if index < len(columns):
                column = columns[index]

            entry[column.lower()] = str(cell.value).split("|")[0] # Values can contains | characters. If they do, we only want to take the first part of the value.

        # Validate entry
        if "address" not in entry:
            raise ValueError("Unable to find 'Address' for entry in excel sheet.")
        if "city" not in entry:
            raise ValueError("Unable to find 'City' for entry in excel sheet.")
        if "start time" not in entry:
            raise ValueError("Unable to find 'Start Time' for entry in excel sheet.")
        if "end time" not in entry:
            raise ValueError("Unable to find 'End Time' for entry in excel sheet.")

        entries.append(entry)

    return RecordSet(columns, entries)

def write_solution(solution, locations, images, record_set, service_time, matrix, path):
    wb = openpyxl.Workbook()

    for i, vehicle in enumerate(solution.vehicles):
        # Create a sheet for the vehicle
        if i == 0:
            ws = wb.active
            ws.title = "Vehicle " + str(i + 1)
        else:
            ws = wb.create_sheet(title = "Vehicle " + str(i + 1))

        # Insert headers
        ws.append(record_set.columns + [ "Arrival Time", "Departure Time", "Travel Time", "Distance"])

        # Append nodes to the worksheet
        for j, node in enumerate(vehicle.nodes):
            if j == 0:  # The first node is the depot. Departure time = Arrival time of next - travel time 
                next_time = vehicle.nodes[1].time
                dist_time = matrix.get_entry(node.location, vehicle.nodes[1].location)
                seconds = next_time - dist_time.time
                departure_time = str(datetime.timedelta(seconds = seconds))
                ws.append(([""] * (len(record_set.columns) + 1)) + [departure_time]) # Empty columns for everything except departure time
            else:
                td = datetime.timedelta(seconds = node.time)
                arrival_time = str(td)
                departure_td = td + datetime.timedelta(minutes = service_time)
                departure_time = str(departure_td)

                prev_location = vehicle.nodes[j - 1].location
                dist_time = matrix.get_entry(prev_location, node.location)
                distance = round(dist_time.distance, 1)

                travel_time = str(datetime.timedelta(seconds = dist_time.time))

                if j == len(vehicle.nodes) - 1: # Arrival at the depot
                    arrival_time = str(datetime.timedelta(seconds = node.time))
                    ws.append(([""] * (len(record_set.columns))) + [arrival_time, "", travel_time, distance])
                else:
                    loc_idx = locations.index(node.location)
                    entry = record_set.entries[loc_idx - 1]
                    row = [entry[k.lower()] or "" for k in record_set.columns]
                    ws.append(row + [ arrival_time, departure_time, travel_time, distance ])

        # Add the map image
        image_index = len(vehicle.nodes) + 1 + 2 # 1 for the headers, 2 as empty space

        img = openpyxl.drawing.image.Image(images[i])
        ws.add_image(img, "A" + str(image_index))

    wb.save(filename = path)